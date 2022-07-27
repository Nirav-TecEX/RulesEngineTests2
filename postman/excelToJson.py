from ast import And
import math
import time
import json
import openpyxl as pyxl
import re
import pandas as pd
import warnings
import sys
import os


# Function reads the testing excel sheet
def read_excel(wbPath, force_url = ""):
    start_time = time.time()

    try:
        with warnings.catch_warnings():
            warnings.filterwarnings(action='ignore', message='Data Validation extension is not supported', category=UserWarning)
            wb = pyxl.load_workbook(wbPath)
            
    except FileNotFoundError:
        print("Could not find file: ", wbPath)
        sys.exit(2)
    except OSError:
        print("Problem reading file: ", wbPath)
        sys.exit(2)

    test_sheet_list = list()
    test_dict_list = list()

    for sheetname in wb.sheetnames:
        if wb[sheetname]['B3'].value is True:
            test_sheet_list.append(wb[sheetname])

    for test_sheet in test_sheet_list:
        test_dict = dict()

        test_dict["name"] = test_sheet['B8'].value
        if not test_dict["name"]:
            print("Error with", wbPath, ":", test_sheet.title, " : No name provided in cell B8. Ignoring sheet.")
            continue

        if not force_url:
            test_dict["url"] = test_sheet['B4'].value
            if not test_dict["url"]:
                print("Error with", wbPath, ":", test_sheet.title, " : No URL provided in cell B4. Ignoring sheet.")
                continue
        else:
            test_dict["url"] = force_url

        keys = []
        input_value = []
        test_key = []
        expected_value = []
        mod = []
        firstTest = 0
        testColumnFound = False
        for row in test_sheet.iter_rows(min_row=0, min_col=1, max_col=2):
            if row[0].value  == "Input Keys":
                print("Found First Test @ : " + str(firstTest))
                break
            elif firstTest >= 250:
                print("Error: 'Input Keys' Not found in Column A, please check if Excel Sheets are Correct")
                break
                
            else: 
                firstTest += 1
            
        
        for row in test_sheet.iter_rows(min_row=firstTest +2, min_col=1, max_col=2):
            if row[0].value is None:
                break
            
           
            if not row[0].value.startswith("#"):
                keys.append(row[0].value)
                #print(str(type(row[1].value)))
                if(str(type(row[1].value)) != "<class 'datetime.datetime'>" ): 
                    input_value.append(row[1].value)   
                else:
                    v = str(row[1].value)
                    v = re.sub("\s", "T",v)
                    v = ''.join((v,".000+0000"))
                    input_value.append(v)  
            
            # Testing Gitbot integration

        for row in test_sheet.iter_rows(min_row=firstTest +2, min_col=1, max_col=5):
            if row[0].value is None or (row[0].value is not None and not row[0].value.startswith("#")):
                # print(type(row[4].value))
                if row[3].value is None  :
                    break
                if row[2].value is True :
                    
                    test_key.append(row[3].value)
                    if(str(type(row[4].value)) == "<class 'datetime.datetime'>" ): 
                        v = str(row[4].value)
                        v = re.sub("\s", "T",v)
                        v = ''.join((v,".000+0000"))
                    else:
                        v = row[4].value
                    if(row[4].value == "!included"):
                        v = "undefined"
                    if isinstance(v, str):
                        if re.search(r'^[0-9]+[.,][0-9]+$', v):
                            v = float(v.replace(",","."))   
                            
                    # if row[0].value is None or not row[0].value.startswith("#"):
                   #
                    expected_value.append(v)
                    mod.append(row[2].value)
                

        test_dict["keys"] = keys
        test_dict["input_value"] = input_value
        test_dict["test_key"] = test_key
        test_dict["expected_value"] = expected_value

        test_dict_list.append(test_dict)

    print("Processed", str(len(test_dict_list)), "sheets in",  "{:.3f}".format(time.time() - start_time), "seconds")

    return test_dict_list

JSONRequestFormat= r"""
  [
    {
      "testName": "Descriptive Test Name",
      "valueTest": [
        {
          "fieldName": "response[0].soId",
          "fieldValue": "request.shipments[0].soId"
        },
        {
          "fieldName": "response[0].productType",
          "fieldValue": "'Tech Equipment'"
        }
      ],
      "numCompare": [
        {
          "fieldName": "'response[0].existingShipmentValueUSD'",
          "operator": ">",
          "value": "1"
        }
      ],
     
      "valueContainTest": [
        {
          "fieldName": "response[0].soId",
          "fieldValue": "'1'"
        }
      ],
      "fieldCheckTest": [
        {
          "fieldName": "response[0].soId"
        },
        {
          "fieldName": "response[0].productType"
        }
      ],
      "requestUrl": "https://8ola05rsaf.execute-api.eu-central-1.amazonaws.com/development/productmatchingrollout",
      "compareUrl": "'https://8ola05rsaf.execute-api.eu-central-1.amazonaws.com/staging-4159b19b/productmatchingrollout'",
      "requestBody": {},
      "expectedResponse" : {}
    }
  ]
"""

# Function duplicates the request payload to the length of the testing sheet
def duplicate_request_payload(test_dict):
#    with open('./JSONRequestFormat.json') as json_file:
#        request_payload = json.load(json_file)
    request_payload = json.loads(JSONRequestFormat)
    for i in range(len(test_dict) - 1):
        request_payload.append(request_payload[0])
    return request_payload


# Function writes the testing excel information to the request payload
def generate_test(workbook_dict,test_name):
    test_json_list = list()

    for one_sheet_dict in workbook_dict:
        
        one_request_entry = dict()
        one_request_entry["testName"] = one_sheet_dict["name"]
        
        one_request_entry["requestUrl"] = one_sheet_dict["url"]
        
        test_list = list()
        
        if len(one_sheet_dict["test_key"]) != len(one_sheet_dict["expected_value"]):
            raise(Exception("Internal Error with length of test_key not equal to length of expected_value lists"))
        for i in range(len(one_sheet_dict["test_key"])):
            new_test = dict()
            new_test['fieldName'] = one_sheet_dict["test_key"][i]
            new_test['fieldValue'] = one_sheet_dict["expected_value"][i]
            test_list.append(new_test)
        one_request_entry["valueTest"] = test_list
       
        request_body_dict = dict()
        if len(one_sheet_dict["keys"]) != len(one_sheet_dict["input_value"]):
            raise(Exception("Internal Error with length of keys not equal to length of input_value lists"))
        for i in range(len(one_sheet_dict["keys"])):
            request_body_dict[one_sheet_dict["keys"][i]] = one_sheet_dict["input_value"][i]
       
        if not "__format_in" in request_body_dict:
            request_body_dict["__format_in"] = "flat"

        if not "__instructions.modules_to_run[0]" in request_body_dict:
            request_body_dict["__instructions.modules_to_run[0]"] = "passthrough_func"

        if not "__instructions.return_path" in request_body_dict:
            request_body_dict["__instructions.return_path"] = ""

        one_request_entry["requestBody"] = request_body_dict
        one_request_entry["requestBody"] = one_request_entry["requestBody"]
        test_json_list.append(one_request_entry)
        
    test_name = os.path.split(test_name)
    json_path = os.path.join(test_name[0], 'jsons', test_name[-1])
    with open(json_path, 'w') as outfile:
        
        outfile.write(json.dumps(test_json_list, indent = 4))
        
        
    #print(json.dumps(request_payload))

    # Writes to Json file
    # with open(f'{excel_name}.json', 'w') as json_payload:
    #     json.dump(request_payload, json_payload)
    # print(f"Writing the test request payload took: {time.time() - start_time}.")


# Changes None / True / False to null / true / false respectively
def null_true_false(element):
    if element is None:
        return "null"
    elif isinstance(element, float) and not math.isfinite(element):
        return "null"
    elif element is True:
        return "true"
    elif element is False:
        return "false"
    elif isinstance(element, dict) and len(element) > 0:
        for k, v in element.items():
            element[k] = null_true_false(v)
    elif isinstance(element, list) and len(element) > 0:
        for i in range(len(element)):
            element[i] = null_true_false(element[i])
    return element


# Removes nan elements from the request payload
def nan_to_none(element):
    if element is None:
        return None
    elif isinstance(element, float) and not math.isfinite(element):
        return None
    elif isinstance(element, dict) and len(element) > 0:
        for k, v in element.items():
            element[k] = nan_to_none(v)
    elif isinstance(element, list) and len(element) > 0:
        for i in range(len(element)):
            element[i] = nan_to_none(element[i])
  
        
    return element

